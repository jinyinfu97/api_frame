# -*- coding=utf-8 -*-
# @time:2021/10/24
# @phone:15874198829
# @author:码尚教育_星瑶
"""
python对excel文件进行操作
一、python对excel文件（.xlsx、.xls）进行操作的库有哪些  第三方库
    xlrd  读excel  xlwt 编辑   xlutils 复制  2010版本以下
    openpyxl  支持2010版本及以上版本excel文件读取及编辑
    pandas   一般用于数据的分析
二、openpyxl模块的基本应用
    1、安装  pip install openpyxl
    2、了解openpyxl模块
    openpyxl模块的三大组件
    1、工作簿
    2、工作表
    3、单元格
三、对excel常用操作有哪些
    1、创建excel文件
    2、读取excel文件数据
    3、编辑excel文件数据

"""
import openpyxl

def read_excel(path):
    """
    读取excel数据   获取工作簿-->找工作表--> 找某个单元格数据
    :param path:
    :return:#操作下将表格中的数据转化为字典列表  [{'id':1,'username':"","passwd":''},{}，{}]
    """
    #获取工作簿
    wk=openpyxl.load_workbook(path)
    #获取工作表
    my_sheet=wk["Sheet1"]
    #获取某个单元格
    value=my_sheet.cell(2,2).value   #B1指坐标
    print(value)
    # 获取工作表中的最大行数及列数
    print("最大行数:",my_sheet.max_row)
    print("最大列数：",my_sheet.max_column)
    datas=[]
    #从excel文件Sheet1工作表一次性读取所有的数据？
    # 循环my_sheet工作表中每一行
    for row in range(2,my_sheet.max_row+1):
        rowdata={}
        #3...n列，3个单元格数据
        #循环每行中的每一列
        for col in range(1,my_sheet.max_column+1):
            key=my_sheet.cell(1,col).value
            value=my_sheet.cell(row,col).value
            rowdata[key]=value
            print(value,end=" ")
        print(rowdata)
        datas.append(rowdata)
        print("")
    print(datas)
    #作业：如何读取工作簿中的所有的数据？
    # 如何获取工作簿中的所有的sheet?
    sheets=wk.sheetnames
    print(sheets)
    return datas

def creat_excel():
    """创建excel文件"""
    #创建工作簿
    wk=openpyxl.Workbook()
    #获取当前工作表(活跃工作表（当前编辑的工作表）)
    my_sheet=wk.active
    #数据写入单元格
    my_sheet.cell(1,1).value="username"
    my_sheet.cell(1, 2).value = "passwd"
    my_sheet.cell(1, 3).value = "adress"
    wk.save('userinfo.xlsx')

def edit_excel(path):
    """编辑excel文件"""
    #加载工作簿
    wk = openpyxl.load_workbook(path)
    #创建新的工作表
    # creat_sheet(wk,"mysheet","userinfo.xlsx")
    # 编辑单元格值
    mysheet=wk["mysheet"]
    mysheet.cell(1,3).value=666
    #编辑后保存
    wk.save(path)

def creat_sheet(wk,sheetname,path):
    my_sheet = wk.create_sheet(sheetname)
    my_sheet.cell(1, 1).value = "1"
    my_sheet.cell(1, 2).value = "2"
    my_sheet.cell(1, 3).value = "3"

def add_manydata(path):
    # 加载工作簿
    wk = openpyxl.load_workbook(path)
    data=[1,2,3,4,5,6]
    mysheet=wk.create_sheet()
    mysheet.append(data)
    wk.save(path)

def read_allsheet(path):
    #字典  {"sheet1":sheet1数据,"sheet2":sheet2数据,...}
    wk=openpyxl.load_workbook(path)
    sheetnames=wk.sheetnames
    alldatas={}
    for sheetname  in sheetnames:
        sheetdatas=[]
        my_sheet=wk[sheetname]
        for row in range(1, my_sheet.max_row + 1):
            rowdata = {}
            for col in range(1, my_sheet.max_column + 1):
                key = my_sheet.cell(1, col).value
                value = my_sheet.cell(row, col).value
                rowdata[key] = value
            sheetdatas.append(rowdata)

        key=sheetname
        value=sheetdatas
        alldatas[key]=value
    print(alldatas)


if __name__ == '__main__':
    # datas=read_excel('info.xlsx')
    # for data in datas:
    #     print(data)
    #     print("用户名",data["username"])
    #     print("密码",data["password"])
    # # creat_excel()  #创建excel文件
    # path="userinfo.xlsx"
    # # edit_excel(path)#编辑excel
    # # add_manydata(path)
    # #内卷
    # read_allsheet('userinfo.xlsx')
    read_excel('../../python212/class16/test_data.xlsx')
