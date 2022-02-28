# 1、创建excel文件，创建userinfo工作表，把[['id','username','passwd'],['01','zhangsan','123456'],['02','lisi','123456'],['03','wangwu','123456'],['04','maliu','123456'],,['05','chenqi','123456']]写入文件中
import os

import openpyxl

def newexcel():
    list0 =[['id','username','passwd'],['01','zhangsan','123456'],['02','lisi','123456'],['03','wangwu','123456'],['04','maliu','123456'],['05','chenqi','123456']]
    # 创建工作簿
    wk = openpyxl.Workbook()
    # 获取当前工作表(活跃工作表（当前编辑的工作表）)
    my_sheet = wk.active
    for i in range(1,7):
        for n in range(1,4):
            my_sheet.cell(i, n).value = list0[i-1][n-1]
    wk.save('userinfo_test.xlsx')
# newexcel()

# 2、一次性把excel表格中所有工作表中所有的数据全部读取，并把数据基于Python数据类型进行存储
# [{"sheetname":"","data":[{"id":"01","username":"zhangsan","passwd":"123456"},.....]]基于这样的格式保存
def excel_info(path):
    # 获取工作簿
    wk = openpyxl.load_workbook(path)

    for num in wk.sheetnames:
        list2 = []
        my_sheet = wk[num]
        # 获取工作表中的最大行数及列数
        print("最大行数:", my_sheet.max_row)
        print("最大列数：", my_sheet.max_column)
        for n in range(2, my_sheet.max_row):
            info = {}
            for i in range(1,my_sheet.max_column + 1):
                key = my_sheet.cell(1, i).value
                value = my_sheet.cell(n, i).value
                info[key]=value
            list2.append(info)
        info1 = {"sheetname":num,"data":list2,'path':os.getcwd()+f'\\{path}'}
        print(info1)
excel_info('userinfo_test.xlsx')
















