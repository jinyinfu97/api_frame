# _*_ coding:utf-8 _*-
# 作者：码尚学院_星瑶老师
# @Time: 2021/5/10
# @Email:2155896749@qq.com
# @File: excel_action.py
import os

import openpyxl

"""
Excel的处理：
1、创建Excel文件

"""

def createExcel():
    """
    创建创建Excel文件，并写入数据
    :return:
    """
    # 创建工作簿
    wk=openpyxl.Workbook()
    # 获取当前工作表
    sheet=wk.active
    # 写数据到单元格
    sheet.cell(1,1).value="username"
    sheet.cell(1, 2).value = "class"
    sheet.cell(1, 3).value = "adress"
    wk.save("userinfo.xlsx")

"""
2、读取Excel文件

"""
def readExcel(filepath):
    # 获取工作簿
    wk=openpyxl.load_workbook(filepath)
    # 方式一：获取工作表
    # sheet1=wk.get_sheet_by_name("Sheet")
    # 方式二：获取工作表
    sheet1=wk["Sheet"]
    # 获取单元格坐标
    location=(sheet1.cell(5,1))
    # 获取单元格值
    value=sheet1.cell(5,1).value
    print(location,value)
    #获取工作表行数及列数
    rows=sheet1.max_row
    cols=sheet1.max_column
    print(f"行数：{rows},列数：{cols}")
    datalist=[]
    #读取工作表中所有数据？
    # 循环excel每一个行
    for row in range(1,rows+1):
        #循环列数，并取值
        for col in range(1,cols+1):
            value=sheet1.cell(row,col).value
            datalist.append(value)
    print(datalist)


def printDataExcel(filepath):
    data=readExcel(filepath)
    print(data)

"""
2、编辑Excel文件
    1、添加一个新的sheet
    2、修改某个单元格数据

"""
def editExcel(filepath):
    # 加载工作簿
    wk=openpyxl.load_workbook(filepath)
    #创建新的工作表
    mysheet=wk.create_sheet("mysheet")
    mysheet.cell(1,1).value="username"
    mysheet.cell(2, 1).value = "DD"
    # 编辑后记得保存
    wk.save(filepath)
    wk.close()


if __name__ == '__main__':
    # createExcel()

    # readExcel(os.getcwd()+r'\userinfo.xlsx')
    editExcel("userinfo.xlsx")






