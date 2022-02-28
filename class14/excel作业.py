# # 1、创建excel文件，插入数据['张三','李四','王伍','马六']
import os
import openpyxl
#
# def createExcel():
#     """
#     创建创建Excel文件，并写入数据
#     :return:
#     """
#     # 创建工作簿
#     wk = openpyxl.Workbook()
#     # 获取当前工作表
#     sheet = wk.active
#     # 写数据到单元格
#     for i in range(0, 4):
#         list = ['张三', '李四', '王伍', '马六']
#         sheet.cell(i + 1, 1).value = list[i]
#     wk.save("userinfo2.xlsx")
# createExcel()
#
# # 2、修改excel文件，修改1行1列的数据为张五
# def editExcel(filepath):
#     # 加载工作簿
#     wk = openpyxl.load_workbook(filepath)
#     # 创建新的工作表
#     sheet = wk.active
#     sheet.cell(1, 1).value = "张五"
#     # 编辑后记得保存
#     wk.save(filepath)
#     wk.close()
# editExcel(os.getcwd() + r"\userinfo2.xlsx")
#
# # 3、新建一个‘MySheet’工作表，用户名分为['小林','修习人生','辉','厨艺天才刘同学']，密码：123456
# # 并且同时插入10相同的数据写入excel文件中
# def editExcel2(filepath):
#     # 加载工作簿
#     wk = openpyxl.load_workbook(filepath)
#     # 创建新的工作表
#     mysheet = wk.create_sheet("mysheet")
#     mysheet.cell(1, 1).value = "username"
#     mysheet.cell(1, 2).value = "password"
#     for i in range(0, 40):
#         list = ['小林', '修习人生', '辉', '厨艺天才刘同学'] * 10
#         mysheet.cell(i + 2, 1).value = list[i]
#     for n in range(0, 40):
#         list1 = ['123456', '123456', '123456', '123456'] * 10
#         mysheet.cell(n + 2, 2).value = list1[n]
#     # 编辑后记得保存
#     wk.save(filepath)
#     wk.close()
# editExcel2(os.getcwd() + r"\userinfo2.xlsx")

# 4、读取‘MySheet’工作表中所有的数据，并且把excel中的数据保存到字典列表中
def readExcel(filepath):
    # 获取工作簿
    wk=openpyxl.load_workbook(filepath)
    # 方式一：获取工作表
    # sheet1=wk.get_sheet_by_name("Sheet")
    # 方式二：获取工作表
    mysheet=wk["mysheet"]
    #获取工作表行数及列数
    rows=mysheet.max_row
    cols=mysheet.max_column
    print(f"行数：{rows},列数：{cols}")
    datalist={}
    #读取工作表中所有数据？
    # 循环excel每一个行
    key = ['用户名','密码']
    newlist = []
    list1 = []
    list2 = []
    for row in range(2,rows+1):
        value = mysheet.cell(row,1).value
        value2 = mysheet.cell(row, 2).value
        list1.append(value)
        list2.append(value2)
    for i in range(len(list1)):
        info = dict.fromkeys(key)
        info['用户名'] = list1[i]
        info['密码'] = list2[i]
        newlist.append(info)
    print(newlist)
    for x in newlist:
        print(x)

readExcel(os.getcwd() + r"\userinfo2.xlsx")



# 5、删除‘MySheet’工作表
